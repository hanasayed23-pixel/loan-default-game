r"""
analyze_session.py
Turns the raw Loan Officer Vignette Experiment data into a deliverable package:
  * Results_LoanOfficer_Experiment.xlsx  - multi-sheet workbook
  * charts/*.png                         - four figures
  * Results_Summary.pdf                  - one-page summary for the supervisor

USAGE (Windows PowerShell)
    py analyze_session.py data\decisions.csv
    py analyze_session.py data\*.csv                 # several session files
    py analyze_session.py data\decisions.csv -o out  # choose output folder
"""

import argparse
import glob
import os
import sys
import textwrap
from datetime import datetime

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import hypothesis_engine as he

NAVY = "#1C2B4A"
ACCENT = "#B08A3E"
GREY = "#8A8F98"
LIGHT = "#EEF1F6"

plt.rcParams.update(
    {
        "font.family": "DejaVu Sans",
        "font.size": 10,
        "axes.edgecolor": "#4A4F58",
        "axes.labelcolor": "#22252A",
        "axes.titlesize": 11,
        "axes.titleweight": "bold",
        "figure.dpi": 150,
        "savefig.dpi": 150,
        "axes.spines.top": False,
        "axes.spines.right": False,
    }
)


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------

def fig_main_effect(df, eff, factor, labels, title, subtitle, effect_col):
    """Bar chart of the two levels plus one line per participant."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9.2, 3.9), gridspec_kw={"width_ratios": [1, 1.15]})

    rate0 = df.loc[df[factor] == 0, "approve"].mean()
    rate1 = df.loc[df[factor] == 1, "approve"].mean()
    ax1.bar([0, 1], [100 * rate0, 100 * rate1], color=[GREY, NAVY], width=0.55)
    for x, v in zip([0, 1], [100 * rate0, 100 * rate1]):
        ax1.text(x, v + 2, f"{v:.0f}%", ha="center", fontweight="bold", color=NAVY)
    ax1.set_xticks([0, 1])
    ax1.set_xticklabels(labels)
    ax1.set_ylabel("Approval rate (%)")
    ax1.set_ylim(0, 108)
    ax1.set_title("Pooled approval rate")

    for _, r in eff.iterrows():
        pid = r["participant_id"]
        g = df[df.participant_id == pid]
        y0 = 100 * g.loc[g[factor] == 0, "approve"].mean()
        y1 = 100 * g.loc[g[factor] == 1, "approve"].mean()
        ax2.plot([0, 1], [y0, y1], marker="o", markersize=5, linewidth=1.6, alpha=0.85, label=f"P{pid}")
    ax2.set_xticks([0, 1])
    ax2.set_xticklabels(labels)
    ax2.set_xlim(-0.25, 1.25)
    ax2.set_ylim(-6, 106)
    ax2.set_ylabel("Approval rate (%)")
    ax2.set_title("Each participant separately")
    ax2.legend(frameon=False, fontsize=8, ncol=2, loc="lower right")

    fig.suptitle(title, fontsize=12, fontweight="bold", color=NAVY, x=0.02, ha="left", y=1.02)
    fig.text(0.02, 0.955, subtitle, fontsize=9, color=GREY, ha="left")
    fig.tight_layout(rect=(0, 0, 1, 0.93))
    return fig


def fig_conflict(h3, eff):
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9.2, 3.9), gridspec_kw={"width_ratios": [1, 1.15]})
    pp = h3["per_participant"]

    a = 100 * pp["approval_A_high_score_nonproductive"].mean()
    b = 100 * pp["approval_B_low_score_productive"].mean()
    ax1.bar([0, 1], [a, b], color=[NAVY, ACCENT], width=0.55)
    for x, v in zip([0, 1], [a, b]):
        ax1.text(x, v + 2, f"{v:.0f}%", ha="center", fontweight="bold", color=NAVY)
    ax1.set_xticks([0, 1])
    ax1.set_xticklabels(["High score\n+ non-productive", "Low score\n+ productive"], fontsize=9)
    ax1.set_ylabel("Approval rate (%)")
    ax1.set_ylim(0, 108)
    ax1.set_title("The two conflict cells")

    x = np.arange(len(eff))
    w = 0.38
    ax2.bar(x - w / 2, 100 * eff["effect_credit_score_H1"], width=w, color=NAVY, label="Credit score (H1)")
    ax2.bar(x + w / 2, 100 * eff["effect_repayment_capacity_H2"], width=w, color=ACCENT, label="Repayment capacity (H2)")
    ax2.axhline(0, color="#4A4F58", linewidth=0.9)
    ax2.set_xticks(x)
    ax2.set_xticklabels([f"P{p}" for p in eff["participant_id"]])
    ax2.set_ylabel("Effect on approval (pp)")
    ax2.set_title("Effect size per participant")
    ax2.legend(frameon=False, fontsize=8)

    fig.suptitle("H3  Does credit score dominate repayment capacity?", fontsize=12,
                 fontweight="bold", color=NAVY, x=0.02, ha="left", y=1.02)
    fig.text(0.02, 0.955,
             f"Conflict-cell difference = {100*h3['statistic']:+.1f} pp, exact p = {h3['p_value']:.4f}",
             fontsize=9, color=GREY, ha="left")
    fig.tight_layout(rect=(0, 0, 1, 0.93))
    return fig


def fig_cells_and_reasons(cells, reasons):
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9.2, 4.1), gridspec_kw={"width_ratios": [1.25, 1]})

    c = cells.copy()
    c["label"] = (
        c["credit_score"].astype(str).str[0]
        + " / "
        + c["use_of_funds"].astype(str).str.slice(0, 8)
        + " / "
        + c["framing"].astype(str).str.slice(0, 7)
    )
    colors = [NAVY if s.lower() == "high" else ACCENT for s in c["credit_score"].astype(str)]
    ax1.barh(range(len(c)), 100 * c["approval_rate"], color=colors, height=0.66)
    ax1.set_yticks(range(len(c)))
    ax1.set_yticklabels(c["label"], fontsize=8)
    ax1.invert_yaxis()
    ax1.set_xlabel("Approval rate (%)")
    ax1.set_xlim(0, 108)
    ax1.set_title("All 8 design cells  (score / use / framing)")
    for i, v in enumerate(100 * c["approval_rate"]):
        ax1.text(v + 1.5, i, f"{v:.0f}%", va="center", fontsize=8, color=NAVY)

    if "Times chosen" in reasons.columns:
        r = reasons.head(6).iloc[::-1]
        ax2.barh(range(len(r)), r["Times chosen"], color=NAVY, height=0.6)
        ax2.set_yticks(range(len(r)))
        ax2.set_yticklabels([str(s)[:26] for s in r.iloc[:, 0]], fontsize=8)
        ax2.set_xlabel("Times chosen")
        ax2.set_title("Self-reported deciding factor")
    else:
        ax2.axis("off")
        ax2.text(0.5, 0.5, "No reason data", ha="center", va="center", color=GREY)

    fig.suptitle("Design cells and stated reasoning", fontsize=12, fontweight="bold",
                 color=NAVY, x=0.02, ha="left", y=1.02)
    fig.tight_layout(rect=(0, 0, 1, 0.94))
    return fig


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

def write_workbook(path, tables, meta):
    thin = Side(style="thin", color="C3C9D4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1C2B4A")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        for sheet, frame in tables.items():
            frame.to_excel(xl, sheet_name=sheet[:31], index=False)
        meta.to_excel(xl, sheet_name="Method_Notes", index=False)

        wb = xl.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
                cell.border = border
            ws.row_dimensions[1].height = 34
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if isinstance(cell.value, float):
                        cell.number_format = "0.000"
            for j, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
                widest = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[get_column_letter(j)].width = min(max(widest + 2, 11), 62)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs", nargs="+", help="decision-level CSV file(s)")
    ap.add_argument("-o", "--outdir", default="results")
    ap.add_argument("--perms", type=int, default=he.N_PERM)
    args = ap.parse_args()

    paths = []
    for pattern in args.inputs:
        hits = glob.glob(pattern)
        paths.extend(hits if hits else [pattern])
    paths = [p for p in paths if os.path.exists(p)]
    if not paths:
        sys.exit("No input file found. Check the path you typed.")

    os.makedirs(args.outdir, exist_ok=True)
    chart_dir = os.path.join(args.outdir, "charts")
    os.makedirs(chart_dir, exist_ok=True)

    df = he.load_and_prepare(paths)
    n_part = df["participant_id"].nunique()
    print(f"Loaded {len(df)} decisions from {n_part} participants across {len(paths)} file(s).")

    eff = he.participant_effects(df)
    cells = he.cell_means(df)
    h1 = he.permutation_test(df, "score_high", n_perm=args.perms)
    h2 = he.permutation_test(df, "use_productive", n_perm=args.perms)
    h3 = he.conflict_cell_test(df)
    lpm = he.lpm_with_wald(df)
    reasons = he.reasons_table(df)
    learning = he.learning_table(df)
    headline = he.hypothesis_table(df, h1, h2, h3, eff, lpm)

    # ---- figures
    f1 = fig_main_effect(df, eff, "score_high", ["Low score", "High score"],
                         "H1  Effect of credit score on approval",
                         f"Within-participant difference = {100*h1['statistic']:+.1f} pp, "
                         f"randomisation p = {h1['p_value']:.4f}",
                         "effect_credit_score_H1")
    f2 = fig_main_effect(df, eff, "use_productive", ["Non-productive", "Productive"],
                         "H2  Effect of prospective repayment capacity on approval",
                         f"Within-participant difference = {100*h2['statistic']:+.1f} pp, "
                         f"randomisation p = {h2['p_value']:.4f}",
                         "effect_repayment_capacity_H2")
    f3 = fig_conflict(h3, eff)
    f4 = fig_cells_and_reasons(cells, reasons)

    names = ["H1_credit_score", "H2_repayment_capacity", "H3_dominance", "cells_and_reasons"]
    for fig, nm in zip([f1, f2, f3, f4], names):
        fig.savefig(os.path.join(chart_dir, f"{nm}.png"), bbox_inches="tight", facecolor="white")

    pdf_path = os.path.join(args.outdir, "Results_Summary.pdf")
    with PdfPages(pdf_path) as pdf:
        cover = plt.figure(figsize=(8.27, 11.69))
        cover.text(0.07, 0.94, "Loan Officer Vignette Experiment", fontsize=19,
                   fontweight="bold", color=NAVY)
        cover.text(0.07, 0.915, "Credit Score or Repayment Capacity?  Pilot session results",
                   fontsize=11, color=GREY)
        cover.text(0.07, 0.893, datetime.now().strftime("Session analysed %d %B %Y, %H:%M"),
                   fontsize=9, color=GREY)

        def block(y, text, size=8.8, color="#22252A", indent=0.07, width=96, weight="normal",
                  family=None, lead=0.0148):
            for line in textwrap.wrap(text, width=width) or [""]:
                kw = {"family": family} if family else {}
                cover.text(indent, y, line, fontsize=size, color=color, fontweight=weight, **kw)
                y -= lead
            return y

        y = 0.858
        cover.text(0.07, y, "Headline findings", fontsize=13, fontweight="bold", color=NAVY)
        y -= 0.026
        for _, r in headline.iterrows():
            cover.text(0.07, y, f"{r['Hypothesis']}    {r['Conclusion']}", fontsize=10.5,
                       fontweight="bold", color=NAVY)
            y -= 0.021
            y = block(y, r["Statement"], size=8.6, indent=0.095, width=92)
            y = block(
                y,
                f"Effect {r['Effect (pp)']:+.1f} pp   |   p = {r['p-value']:.4f}   |   "
                f"predicted direction in {r['Participants showing predicted direction']} participants",
                size=8.6, color=GREY, indent=0.095, width=92,
            )
            y -= 0.016

        y -= 0.008
        cover.text(0.07, y, "Design and sample", fontsize=13, fontweight="bold", color=NAVY)
        y -= 0.026
        for line in [
            "2 x 2 x 2 within-subject factorial: credit score (high / low) x use of funds "
            "(productive / non-productive) x urgency framing (planned / urgent). All eight cells "
            "shown once per participant in randomised order.",
            f"{n_part} participants x 8 vignettes = {len(df)} decisions. Every hard financial field "
            "(loan amount, monthly income, years in business, guarantor, field investigation, "
            "literacy) held identical across all vignettes; only the credit score and the one-line "
            "stated purpose vary.",
            "True default probability is set by credit score alone (high 20%, low 45%) and is "
            "independent of the stated purpose. Any weight placed on the purpose is therefore, by "
            "construction, economically non-diagnostic.",
            f"Inference uses randomisation tests that permute factor labels within participant. "
            f"With {n_part} participants this is a demonstration pilot: the direction of each effect "
            "and its consistency across participants carry the interpretation, not conventional "
            "significance thresholds.",
        ]:
            y = block(y, line, size=8.6, width=98)
            y -= 0.006

        y -= 0.010
        cover.text(0.07, y, "Supporting model (linear probability, participant fixed effects)",
                   fontsize=13, fontweight="bold", color=NAVY)
        y -= 0.026
        cs = lpm["coefficients"]
        for term, pretty in [
            ("credit_score_high_H1", "Credit score high  (H1)"),
            ("use_productive_H2", "Productive use     (H2)"),
            ("framing_urgent", "Urgent framing        "),
        ]:
            row = cs[cs.term == term].iloc[0]
            cover.text(0.075, y, f"{pretty}   b = {row['coefficient']:+.3f}   "
                                 f"(se {row['std_error']:.3f})",
                       fontsize=8.8, color="#22252A", family="DejaVu Sans Mono")
            y -= 0.016
        y -= 0.006
        y = block(y, f"Wald test of the H3 restriction (b_score = b_capacity):  "
                     f"F(1, {lpm['dof']}) = {lpm['wald_statistic']:.3f},  p = {lpm['wald_p_value']:.4f}. "
                     f"Reported for continuity with the proposal; treat the exact conflict-cell "
                     f"randomisation test above as the primary evidence on H3.",
                  size=8.6, width=98)
        cover.text(0.07, 0.045, "Generated by analyze_session.py   |   full tables in "
                                "Results_LoanOfficer_Experiment.xlsx", fontsize=7.6, color=GREY)
        pdf.savefig(cover, bbox_inches="tight")
        plt.close(cover)

        for fig in [f1, f2, f3, f4]:
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

    # ---- workbook
    meta = pd.DataFrame(
        {
            "Item": [
                "Design",
                "Participants",
                "Decisions",
                "Dependent variable",
                "H1 test",
                "H2 test",
                "H3 test (primary)",
                "H3 test (secondary)",
                "Why not mixed-effects logit",
                "True default probability",
                "Permutation draws",
                "Generated",
                "Source of every number in this workbook",
            ],
            "Detail": [
                "2x2x2 within-subject factorial: credit score x use of funds x urgency framing; "
                "all eight cells shown once per participant in randomised order.",
                n_part,
                len(df),
                "Binary approve/reject. Secondary continuous measure: approved amount in EGP.",
                "Monte-Carlo randomisation test, labels permuted within participant, "
                f"{args.perms} draws, seed {he.RNG_SEED}.",
                f"Monte-Carlo randomisation test, same procedure, {args.perms} draws.",
                "Exact randomisation test by full enumeration on the two conflict cells "
                f"({h3['n_enumerated']} enumerated relabelings).",
                "Wald test of the equality restriction in a linear probability model with "
                "participant fixed effects, reported for continuity with the proposal.",
                f"With {n_part} clusters, asymptotic cluster-robust and mixed-effects standard "
                "errors are not reliable; randomisation inference is exact under the sharp null.",
                "High credit score 20%, low credit score 45%, independent of stated purpose.",
                args.perms,
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                "Computed directly from the raw decision rows in the Raw_Data sheet by "
                "analyze_session.py; no figure is entered by hand.",
            ],
        }
    )

    tables = {
        "Hypotheses_H1_H2_H3": headline,
        "By_Participant": eff.round(3),
        "Cell_Means_8_cells": cells.round(3),
        "H3_Conflict_Cells": h3["per_participant"].round(3),
        "Model_Coefficients": lpm["coefficients"].round(4),
        "Reasons_Given": reasons,
        "Feedback_Learning_Check": learning,
        "Raw_Data": df,
    }
    xlsx_path = os.path.join(args.outdir, "Results_LoanOfficer_Experiment.xlsx")
    write_workbook(xlsx_path, tables, meta)

    print("\n" + "=" * 68)
    print(headline[["Hypothesis", "Effect (pp)", "p-value", "Conclusion"]].to_string(index=False))
    print("=" * 68)
    print(f"\nWorkbook : {xlsx_path}")
    print(f"Summary  : {pdf_path}")
    print(f"Charts   : {chart_dir}")


if __name__ == "__main__":
    main()
